"""
Comparador Cecotec v4
· Matching 100% local (pandas) — sin IA, sin scraping, sin límites
· Carga feed oficial Cecotec + Keepa exports
· Procesa TODOS los productos Keepa relevantes en segundos
"""

import streamlit as st
import pandas as pd
import requests
from bs4 import BeautifulSoup
import re, time, os
from pathlib import Path
import google.generativeai as genai

st.set_page_config(page_title="Comparador Cecotec", page_icon="🔍", layout="wide")

# ── CSS corporativo ───────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Nunito:wght@400;600;700;800;900&family=Nunito+Sans:wght@400;600;700&display=swap');
:root{--cec-blue:#3EB1C8;--cec-blue-d:#2a8fa3;--cec-black:#141413;--cec-bg:#FAF9F5;--cec-white:#fff;--cec-grey:#e8e8e4;--cec-muted:#6b7280;}
html,body,[class*="css"]{font-family:'Nunito Sans',sans-serif;background:var(--cec-bg)!important;color:var(--cec-black);}
.stApp{background:var(--cec-bg)!important;}
.cec-navbar{background:var(--cec-black);padding:.7rem 2rem;display:flex;align-items:center;gap:1.2rem;margin:-1rem -1rem 1.8rem -1rem;border-bottom:3px solid var(--cec-blue);}
.cec-navbar .logo-text{font-family:'Nunito',sans-serif;font-size:1.55rem;font-weight:900;color:#fff;letter-spacing:-.5px;}
.cec-navbar .logo-text span{color:var(--cec-blue);}
.cec-navbar .logo-sub{font-size:.72rem;color:rgba(255,255,255,.55);font-weight:600;text-transform:uppercase;letter-spacing:.12em;border-left:1px solid rgba(255,255,255,.2);padding-left:1.1rem;}
.cec-navbar .logo-badge{margin-left:auto;background:var(--cec-blue);color:var(--cec-black);font-size:.68rem;font-weight:800;padding:3px 10px;border-radius:20px;text-transform:uppercase;letter-spacing:.08em;}
.cec-section-title{font-family:'Nunito',sans-serif;font-size:1rem;font-weight:800;color:var(--cec-black);text-transform:uppercase;letter-spacing:.08em;padding:.4rem 0 .4rem .7rem;border-left:4px solid var(--cec-blue);margin:1.4rem 0 .9rem;background:linear-gradient(90deg,rgba(62,177,200,.07) 0%,transparent 100%);}
.kpi-row{display:flex;gap:.8rem;margin-bottom:1.6rem;flex-wrap:wrap;}
.kpi{background:var(--cec-white);border-radius:10px;padding:1rem 1.3rem;box-shadow:0 1px 3px rgba(20,20,19,.08),0 0 0 1px var(--cec-grey);flex:1;min-width:120px;position:relative;overflow:hidden;}
.kpi::after{content:'';position:absolute;bottom:0;left:0;right:0;height:3px;background:var(--cec-blue);}
.kpi .val{font-size:2rem;font-weight:900;color:var(--cec-blue);font-family:'Nunito',sans-serif;line-height:1;}
.kpi .lbl{font-size:.68rem;color:var(--cec-muted);text-transform:uppercase;letter-spacing:.07em;margin-top:.3rem;font-weight:600;}
.tag-mejor{background:rgba(62,177,200,.12);color:#0a6a7a;padding:3px 12px;border-radius:4px;font-size:.76rem;font-weight:800;border:1px solid rgba(62,177,200,.35);text-transform:uppercase;}
.tag-igual{background:rgba(255,193,7,.12);color:#856404;padding:3px 12px;border-radius:4px;font-size:.76rem;font-weight:800;border:1px solid rgba(255,193,7,.35);text-transform:uppercase;}
.tag-peor{background:rgba(220,53,69,.1);color:#842029;padding:3px 12px;border-radius:4px;font-size:.76rem;font-weight:800;border:1px solid rgba(220,53,69,.3);text-transform:uppercase;}
.tag-skip{background:var(--cec-grey);color:var(--cec-muted);padding:3px 12px;border-radius:4px;font-size:.76rem;font-weight:800;text-transform:uppercase;}
div[data-testid="stTabs"] button[aria-selected="true"]{color:var(--cec-blue)!important;border-bottom-color:var(--cec-blue)!important;}
div[data-testid="stButton"] button[kind="primary"]{background:var(--cec-blue)!important;border-color:var(--cec-blue)!important;color:#fff!important;font-weight:800;border-radius:6px;}
div[data-testid="stButton"] button[kind="primary"]:hover{background:var(--cec-blue-d)!important;}
div[data-testid="stDataFrame"] thead th{background:var(--cec-black)!important;color:#fff!important;font-weight:700;font-size:.78rem;text-transform:uppercase;}
.stProgress > div > div{background:var(--cec-blue)!important;}
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="cec-navbar">
  <div class="logo-text">ceco<span>tec</span></div>
  <div class="logo-sub">Comparador de competencia</div>
  <div class="logo-badge">⚡ Matching local</div>
</div>
""", unsafe_allow_html=True)

# ── Constantes ────────────────────────────────────────────────────────────────
KEEPA_HOGAR   = "KeepaExport-2026-05-19-BestSellersList-9-599391031.xlsx"
KEEPA_BELLEZA = "BellezaKeepaExport-2026-05-19-BestSellersList-9-4347698031-9000.xlsx"
FEED_FILE     = "feed_Espan_a.xlsx"
STOCK_FILE    = "stock_cecotec.csv"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0 Safari/537.36",
    "Accept-Language": "es-ES,es;q=0.9",
}

# Mapa de categorías Keepa → categorías Cecotec (puede ser lista para cubrir variantes)
# CAT_MAP: keepa_subcat → (cecotec_cats, required_keywords, excluded_keywords)
# required_kw: al menos una debe aparecer en título Cecotec (vacío = sin restricción)
# excluded_kw: ninguna debe aparecer en título (filtra repuestos, accesorios, etc.)
CAT_MAP = {
    # ── Aspiración ────────────────────────────────────────────────────────────
    "aspiradoras escoba":            (["aspiradores verticales"], [], ["repuesto","batería","filtro","accesorio"]),
    "aspiradoras verticales":        (["aspiradores verticales"], [], ["repuesto","batería","filtro"]),
    "aspiradoras de mano":           (["aspiradores de mano"], [], ["repuesto","batería","filtro"]),
    "robots aspiradores":            (["robots aspiradores"], [], ["repuesto","filtro","cepillo lateral","mopa"]),
    "aspiradoras para alfombras":    (["aspiradores de trineo"], [], ["repuesto","bolsa","filtro"]),
    "aspiradoras con bolsa":         (["aspiradores de trineo"], [], ["repuesto","bolsa","filtro"]),
    # ── Planchado ─────────────────────────────────────────────────────────────
    # Plancha horizontal ≠ plancha vertical/cepillo vapor
    "planchas de vapor":             (["planchas de vapor","centro de planchado"], [], ["vertical","viaje","vaporeta","repuesto","suela","depósito","filtro"]),
    "centros de planchado":          (["centro de planchado"], [], ["repuesto","funda","tabla"]),
    "planchas de vapor verticales para viaje": (["planchas verticales","vaporetas"], [], ["repuesto","suela","depósito","filtro"]),
    "cepillos de vapor":             (["vaporetas","planchas verticales"], [], ["repuesto","suela","filtro"]),
    # ── Cocina ────────────────────────────────────────────────────────────────
    "freidoras de aire":             (["freidoras sin aceite"], [], ["repuesto","accesorio","bandeja","molde","papel","bolsa"]),
    "freidoras":                     (["freidoras sin aceite"], [], ["repuesto","accesorio","papel","bolsa"]),
    "hornos de sobremesa":           (["microondas de sobremesa"], [], ["repuesto","plato giratorio","bandeja"]),
    "tostadoras":                    (["tostadoras"], [], ["repuesto"]),
    "sandwicheras":                  (["sandwicheras","grills"], [], ["repuesto","plancha recambio"]),
    "grills de contacto":            (["grills","sandwicheras"], [], ["repuesto"]),
    "batidoras de mano":             (["batidoras de mano"], [], ["repuesto","accesorio","vaso","pie"]),
    "batidoras de vaso":             (["batidoras de vaso"], [], ["repuesto","jarra","accesorio","cuchilla"]),
    "procesadores de alimentos":     (["batidoras / picadoras","robots de cocina"], [], ["repuesto","cuchilla","accesorio"]),
    "batidoras amasadoras":          (["amasadoras","robots de cocina"], [], ["repuesto","accesorio","gancho"]),
    "cafeteras italianas":           (["cafeteras express"], [], ["repuesto","cápsula","accesorio","filtro","descalcificador"]),
    "cafeteras de filtro":           (["cafeteras de filtro"], [], ["repuesto","filtro","jarra","descalcificador"]),
    "cafeteras espresso":            (["cafeteras express"], [], ["repuesto","cápsula","accesorio"]),
    "máquinas de café":              (["cafeteras express"], [], ["repuesto","cápsula"]),
    "cafeteras individuales":        (["cafeteras express"], [], ["repuesto","cápsula"]),
    "hervidores":                    (["hervidores"], [], ["repuesto","filtro"]),
    "robots de cocina":              (["robots de cocina"], [], ["repuesto","accesorio","cuchilla","vaso"]),
    "sartenes para freír":           (["sartenes"], [], ["repuesto","mango","tapa"]),
    "juegos de sartenes":            (["sartenes"], [], ["repuesto"]),
    # ── Básculas ──────────────────────────────────────────────────────────────
    "balanzas digitales":            (["básculas de cocina"], [], ["repuesto"]),
    "básculas de cocina":            (["básculas de cocina"], [], ["repuesto"]),
    "básculas de baño":              (["básculas de baño"], [], ["repuesto"]),
    # ── Clima ─────────────────────────────────────────────────────────────────
    "purificadores de aire":         (["purificadores de aire"], [], ["repuesto","filtro hepa","accesorio"]),
    "humidificadores":               (["humidificadores"], [], ["repuesto","filtro"]),
    "ventiladores":                  (["ventiladores de pie","ventiladores de techo"], [], ["repuesto","mando","control"]),
    "ventiladores de techo":         (["ventiladores de techo"], [], ["repuesto","aspa","mando"]),
    "aires acondicionados portátiles":(["aires acondicionados"], [], ["repuesto","filtro","mando"]),
    "deshumidificadores":            (["deshumidificadores"], [], ["repuesto","filtro"]),
    # ── Audio/Video ───────────────────────────────────────────────────────────
    "televisores":                   (["televisores / smart tv"], [], ["repuesto","soporte","mando","protector"]),
    "monitores":                     (["monitores"], [], ["repuesto","soporte","brazo"]),
    "altavoces portátiles":          (["altavoces"], [], ["repuesto"]),
    # ── Gran electrodoméstico ─────────────────────────────────────────────────
    "lavadoras":                     (["lavadoras"], [], ["repuesto","tambor","correa","bomba","goma"]),
    "lavavajillas":                  (["lavavajillas"], [], ["repuesto","cesta","portavasos"]),
    "frigoríficos":                  (["frigoríficos combi","frigoríficos americanos"], [], ["repuesto","balda","cajón","junta"]),
    "minibar":                       (["minibar / mini nevera"], [], ["repuesto"]),
    "campanas extractoras":          (["campanas extractoras"], [], ["repuesto","filtro","carbón activo"]),
    "microondas sencillos":          (["microondas de sobremesa"], [], ["repuesto","plato giratorio"]),
    "hornos":                        (["hornos integrables"], [], ["repuesto","bandeja","rejilla"]),
    "vinotecas":                     (["vinoteca"], [], ["repuesto"]),
    # ── Belleza ───────────────────────────────────────────────────────────────
    "secadores de pelo":             (["secadores de pelo"], [], ["repuesto","difusor","boquilla","accesorio"]),
    "planchas para el pelo":         (["planchas de pelo"], [], ["repuesto","accesorio"]),
    "planchas de pelo":              (["planchas de pelo"], [], ["repuesto","accesorio"]),
    "rizadores":                     (["rizadores"], [], ["repuesto","accesorio"]),
    "cepillos eléctricos para el cabello": (["cepillos alisadores","planchas de pelo"], [], ["repuesto"]),
    "afeitadoras eléctricas":        (["afeitadoras","depilación"], [], ["repuesto","cabezal","lámina","accesorio"]),
    "depiladores":                   (["depilación","depiladores"], [], ["repuesto","cabezal","accesorio"]),
    "cepillos de dientes eléctricos":(["cepillos de dientes"], [], ["repuesto","cabezal recambio","funda"]),
    "masajeadores":                  (["masajeadores"], [], ["repuesto","accesorio"]),
    "freidoras sin aceite":          (["freidoras sin aceite"], [], ["repuesto","accesorio","papel","bandeja"]),
}

# ── Carga feed Cecotec ────────────────────────────────────────────────────────
@st.cache_data
def load_cecotec_feed(path_str: str) -> pd.DataFrame:
    p = Path(path_str.rstrip("/"))
    path = p if p.suffix in (".xlsx",".xls") else p / FEED_FILE
    if not path.exists():
        return pd.DataFrame()
    df = pd.read_excel(path)
    def strip_html(s):
        return re.sub(r"<[^>]+>", " ", str(s or "")).strip()
    df["desc_clean"] = df["alternate_description"].apply(strip_html)
    df["desc_clean"] = df["desc_clean"].where(df["desc_clean"].str.len() > 10, df["title"].astype(str))
    df_stock = df[
        (df["availability"] == "in stock") &
        (~df["categories"].str.lower().str.contains("repuesto|recambio", na=False))
    ].copy()
    df_stock["price"]       = pd.to_numeric(df_stock["price"], errors="coerce")
    df_stock["sale_price"]  = pd.to_numeric(df_stock["sale_price"], errors="coerce")
    df_stock["precio_final"]= df_stock["sale_price"].fillna(df_stock["price"])
    df_stock["cat_lower"]   = df_stock["categories"].str.lower().fillna("")
    df_stock["title_lower"] = df_stock["title"].str.lower().fillna("")
    df_stock["desc_lower"]  = df_stock["desc_clean"].str.lower().fillna("")
    return df_stock.reset_index(drop=True)


# ── Carga stock Cecotec ───────────────────────────────────────────────────────
@st.cache_data
def load_stock(path_str: str) -> pd.DataFrame:
    """Busca stock_cecotec*.csv en el directorio o acepta ruta directa al fichero."""
    import glob
    p = Path(path_str.rstrip("/"))
    if p.suffix == ".csv" and p.exists():
        candidates = [p]
    else:
        candidates = sorted(glob.glob(str(p / "stock_cecotec*.csv")))
    if not candidates:
        return pd.DataFrame()
    df = pd.read_csv(str(candidates[0]), encoding="latin-1", sep=";", quoting=3, on_bad_lines="skip")
    for col in ["Stock Operativo", "Mar", "Puerto"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)
    df["Referencia"] = df["Referencia"].astype(str).str.strip()
    return df[["Referencia", "Stock Operativo", "Mar", "Puerto"]].drop_duplicates("Referencia")

# ── Carga Keepa (sin límite de 100) ──────────────────────────────────────────
@st.cache_data
def load_keepa_files(upload_dir: str) -> pd.DataFrame:
    dfs = []
    for fname, source in [(KEEPA_HOGAR,"Hogar"),(KEEPA_BELLEZA,"Belleza")]:
        p = Path(upload_dir) / fname
        if not p.exists(): continue
        df = pd.read_excel(p)
        df["_source"] = source
        dfs.append(df)
    if not dfs: return pd.DataFrame()
    return _process_keepa(pd.concat(dfs, ignore_index=True))

def load_custom_file(file) -> pd.DataFrame:
    df = pd.read_csv(file) if file.name.endswith(".csv") else pd.read_excel(file)
    df["_source"] = "Fichero propio"
    return _process_keepa(df)

def _process_keepa(df: pd.DataFrame) -> pd.DataFrame:
    df = df.rename(columns={
        "ASIN":"asin","Título":"titulo","Caja de Compra: Actual":"precio",
        "Categorías: Principal":"categoria_principal","Categorías: Subcategoría":"subcategoria",
        "Clasificación de Ventas: Actual":"ranking","Fabricante":"fabricante",
        "URL: Amazon":"url_amazon","Códigos de producto: EAN":"ean",
        "Opiniones: Valoraciones":"rating","Opiniones: Cantidad de valoraciones":"num_reviews",
        "Descripción & Características: Característica 1":"feat1",
        "Descripción & Características: Característica 2":"feat2",
        "Descripción & Características: Característica 3":"feat3",
        "Descripción & Características: Característica 4":"feat4",
        "Descripción & Características: Descripción breve":"descripcion_breve",
    })
    for col in ["asin","titulo","precio","subcategoria","ranking","fabricante",
                "url_amazon","feat1","feat2","feat3","feat4","descripcion_breve"]:
        if col not in df.columns: df[col] = ""
    # Excluir productos propios Cecotec (solo interesan competidores)
    if "fabricante" in df.columns:
        df = df[~df["fabricante"].str.lower().str.contains("cecotec", na=False)].copy()
    if "ranking" in df.columns:
        df = df.sort_values("ranking")
    # No limit — process ALL products
    df["_cecotec_relevant"] = df["subcategoria"].apply(
        lambda s: str(s).lower().strip() in CAT_MAP
    )
    def build_feats(row):
        parts = [str(row.get(f"feat{i}","") or "") for i in range(1,5)]
        parts.append(str(row.get("descripcion_breve","") or ""))
        return " | ".join(p[:120] for p in parts if p.strip())[:500]
    df["caracteristicas"] = df.apply(build_feats, axis=1)
    return df

# ── Matching local 100% pandas ────────────────────────────────────────────────
def _extract_specs(text: str) -> dict:
    """Extract numeric specs from a description string."""
    text = text.lower()
    specs = {}
    # Watts
    m = re.search(r'(\d[\d.,]+)\s*w\b', text)
    if m: specs["w"] = float(m.group(1).replace(",","."))
    # Pressure Pa/kPa
    m = re.search(r'(\d[\d.,]+)\s*kpa', text)
    if m: specs["kpa"] = float(m.group(1).replace(",","."))
    m = re.search(r'(\d[\d.,]+)\s*pa\b', text)
    if m: specs["pa"] = float(m.group(1).replace(",","."))
    # Steam g/min
    m = re.search(r'(\d[\d.,]+)\s*g/min', text)
    if m: specs["g_min"] = float(m.group(1).replace(",","."))
    # Capacity L
    m = re.search(r'(\d[\d.,]+)\s*l\b', text)
    if m: specs["litros"] = float(m.group(1).replace(",","."))
    # Autonomy min
    m = re.search(r'(\d+)\s*min', text)
    if m: specs["minutos"] = int(m.group(1))
    # Temperature
    m = re.search(r'(\d+)\s*º|(\d+)\s*grados', text)
    if m: specs["temp"] = int((m.group(1) or m.group(2)))
    return specs

def _build_conclusion(ref: dict, row, precio_ref: float, precio_cec: float, prestaciones: str) -> str:
    """Generate a human-readable conclusion explaining the verdict."""
    ahorro = precio_ref - precio_cec
    pct    = (ahorro / precio_ref * 100) if precio_ref > 0 else 0

    ref_text = (str(ref.get("titulo","")) + " " + str(ref.get("caracteristicas",""))).lower()
    cec_text = (str(row["title"]) + " " + str(row["desc_clean"])).lower()

    ref_specs = _extract_specs(ref_text)
    cec_specs = _extract_specs(cec_text)

    puntos_favor    = []
    puntos_contra   = []
    puntos_neutros  = []

    # ── Precio ────────────────────────────────────────────────────────────────
    if ahorro > 0:
        puntos_favor.append(f"precio {ahorro:.2f}€ más barato ({pct:.0f}% de ahorro)")
    elif ahorro < 0:
        puntos_contra.append(f"precio {abs(ahorro):.2f}€ más caro que el competidor")

    # ── Potencia W ────────────────────────────────────────────────────────────
    if "w" in ref_specs and "w" in cec_specs:
        diff_w = cec_specs["w"] - ref_specs["w"]
        if diff_w >= 200:
            puntos_favor.append(f"mayor potencia ({int(cec_specs['w'])}W vs {int(ref_specs['w'])}W)")
        elif diff_w <= -200:
            puntos_contra.append(f"menor potencia ({int(cec_specs['w'])}W vs {int(ref_specs['w'])}W)")
        else:
            puntos_neutros.append(f"potencia similar ({int(cec_specs['w'])}W)")

    # ── Presión kPa/Pa ─────────────────────────────────────────────────────
    for key, label in [("kpa","kPa"), ("pa","Pa")]:
        if key in ref_specs and key in cec_specs:
            diff = cec_specs[key] - ref_specs[key]
            if diff >= ref_specs[key] * 0.1:
                puntos_favor.append(f"mayor succión ({cec_specs[key]:.0f} {label} vs {ref_specs[key]:.0f})")
            elif diff <= -ref_specs[key] * 0.1:
                puntos_contra.append(f"menor succión ({cec_specs[key]:.0f} {label} vs {ref_specs[key]:.0f})")

    # ── Vapor g/min ────────────────────────────────────────────────────────
    if "g_min" in ref_specs and "g_min" in cec_specs:
        diff = cec_specs["g_min"] - ref_specs["g_min"]
        if diff >= 5:
            puntos_favor.append(f"más vapor ({cec_specs['g_min']:.0f} g/min vs {ref_specs['g_min']:.0f})")
        elif diff <= -5:
            puntos_contra.append(f"menos vapor ({cec_specs['g_min']:.0f} g/min vs {ref_specs['g_min']:.0f})")
        else:
            puntos_neutros.append(f"vapor similar ({cec_specs['g_min']:.0f} g/min)")

    # ── Capacidad ─────────────────────────────────────────────────────────
    if "litros" in ref_specs and "litros" in cec_specs:
        diff = cec_specs["litros"] - ref_specs["litros"]
        if diff >= 0.3:
            puntos_favor.append(f"mayor capacidad ({cec_specs['litros']:.1f}L vs {ref_specs['litros']:.1f}L)")
        elif diff <= -0.3:
            puntos_contra.append(f"menor capacidad ({cec_specs['litros']:.1f}L vs {ref_specs['litros']:.1f}L)")

    # ── Autonomía ──────────────────────────────────────────────────────────
    if "minutos" in ref_specs and "minutos" in cec_specs:
        diff = cec_specs["minutos"] - ref_specs["minutos"]
        if diff >= 10:
            puntos_favor.append(f"más autonomía ({cec_specs['minutos']} min vs {ref_specs['minutos']} min)")
        elif diff <= -10:
            puntos_contra.append(f"menos autonomía ({cec_specs['minutos']} min vs {ref_specs['minutos']} min)")

    # ── Funcionalidades clave ──────────────────────────────────────────────
    features_check = [
        (["wifi","connected","app","smart"], "conectividad WiFi/App"),
        (["hepa","h13","h14"], "filtro HEPA"),
        (["autovacío","auto-vaciado","vaciado automático"], "vaciado automático"),
        (["display","pantalla","lcd"], "pantalla/display"),
        (["inox","acero inoxidable"], "acabado inox"),
        (["sin bolsa","sin cable","inalámbric"], "sin cable/bolsa"),
        (["golpe vapor","boost","turbo"], "función turbo/golpe vapor"),
        (["doble voltaje","voltaje universal"], "doble voltaje"),
    ]
    for kws, label in features_check:
        ref_has = any(k in ref_text for k in kws)
        cec_has = any(k in cec_text for k in kws)
        if ref_has and cec_has:
            puntos_neutros.append(f"ambos con {label}")
        elif not ref_has and cec_has:
            puntos_favor.append(f"incorpora {label} (el competidor no)")
        elif ref_has and not cec_has:
            puntos_contra.append(f"sin {label} (el competidor sí lo tiene)")

    # ── Categoría diferente ────────────────────────────────────────────────
    ref_subcat = str(ref.get("subcategoria","")).lower()
    cec_cat    = str(row["categories"]).lower()
    CAT_DIFF_NOTES = {
        ("cepillos de vapor", "vaporeta"): "es una vaporeta de mano, no un cepillo de vapor específico",
        ("planchas de vapor verticales para viaje", "planchas verticales"): "plancha vertical, función similar para viaje",
        ("procesadores de alimentos", "batidoras"): "batidora/picadora, funciones similares aunque sin todos los accesorios",
    }
    for (ref_k, cec_k), nota in CAT_DIFF_NOTES.items():
        if ref_k in ref_subcat and cec_k in cec_cat:
            puntos_neutros.append(nota)

    # ── Construir texto final ──────────────────────────────────────────────
    partes = []
    if prestaciones == "mejor":
        partes.append("**✅ Cecotec es mejor opción** porque")
        items = puntos_favor[:3]
        if puntos_contra:
            items_contra = puntos_contra[:1]
        else:
            items_contra = []
    elif prestaciones == "peor":
        partes.append("**🔴 Cecotec es inferior** ya que")
        items = puntos_contra[:3]
        items_contra = []
    else:
        partes.append("**🟡 Prestaciones equivalentes**:")
        items = puntos_favor[:2] + puntos_neutros[:2]
        items_contra = puntos_contra[:1]

    if items:
        partes.append(", ".join(items))
    if items_contra:
        partes.append(f"aunque {', '.join(items_contra)}")
    if puntos_neutros and prestaciones == "igual" and not items:
        partes.append(", ".join(puntos_neutros[:2]))

    conclusion = " ".join(partes)
    if not conclusion.strip() or conclusion.strip() in ("**✅ Cecotec es mejor opción** porque", "**🟡 Prestaciones equivalentes**:", "**🔴 Cecotec es inferior** ya que"):
        # Fallback genérico
        if prestaciones == "mejor":
            conclusion = f"**✅ Mejor opción:** precio {ahorro:.2f}€ más barato ({pct:.0f}% ahorro) en la misma categoría"
        elif prestaciones == "peor":
            conclusion = f"**🔴 Precio mayor** que el competidor en {abs(ahorro):.2f}€, pero Cecotec con garantía y servicio directo"
        else:
            conclusion = f"**🟡 Alternativa equivalente** a menor precio ({ahorro:.2f}€ de ahorro)"

    return conclusion

def _make_alt(row, precio_ref, ref=None):
    precio_cec = float(row["precio_final"])
    ahorro = round(precio_ref - precio_cec, 2) if precio_ref > 0 else 0.0
    prestaciones = "igual"
    if ahorro > precio_ref * 0.25:
        prestaciones = "mejor"
    elif precio_cec > precio_ref:
        prestaciones = "peor"
    conclusion = _build_conclusion(ref or {}, row, precio_ref, precio_cec, prestaciones)
    return {
        "cecotec_nombre":         row["title"],
        "cecotec_precio":         precio_cec,
        "cecotec_precio_original": float(row["price"]) if row["price"] != row["precio_final"] else None,
        "cecotec_caracteristicas": row["desc_clean"][:200],
        "cecotec_url":            row["link"],
        "cecotec_referencia":     str(row.get("mpn","") or ""),
        "cecotec_stock":          True,
        "cecotec_categoria":      row["categories"],
        "cecotec_imagen":         row.get("image_link",""),
        "ahorro_eur":             ahorro,
        "prestaciones":           prestaciones,
        "justificacion":          conclusion,
    }

def find_best_match_local(ref: dict, df_cec: pd.DataFrame) -> dict:
    precio_ref = float(ref.get("precio") or 0)
    subcat_ref = str(ref.get("subcategoria","")).lower().strip()
    titulo_ref = str(ref.get("titulo","")).lower()
    feats_ref  = str(ref.get("caracteristicas","")).lower()
    texto_ref  = titulo_ref + " " + feats_ref

    # 1. Obtener categorías + filtros del mapa
    map_entry = CAT_MAP.get(subcat_ref)
    if map_entry:
        cec_cats, req_kw, exc_kw = map_entry
    else:
        # Fallback fuzzy
        words = re.findall(r'\w{5,}', subcat_ref)
        cec_cats = [c for c in df_cec["cat_lower"].unique() if any(w in c for w in words)][:3]
        req_kw, exc_kw = [], []

    if not cec_cats:
        return {"no_encontrado": True, "motivo": f"Categoría '{subcat_ref}' sin equivalente en Cecotec"}

    # 2. Filtrar por categoría
    mask_cat = df_cec["cat_lower"].apply(lambda c: any(cc.lower() in c for cc in cec_cats))
    df_f = df_cec[mask_cat].copy()
    if df_f.empty:
        return {"no_encontrado": True, "motivo": f"Sin productos en: {', '.join(cec_cats)}"}

    # 3. Aplicar exclusiones (repuestos, accesorios, etc.)
    if exc_kw:
        exc_pattern = "|".join(exc_kw)
        df_f = df_f[~df_f["title_lower"].str.contains(exc_pattern, na=False)]

    # 4. Aplicar keywords requeridos (filtro de tipo de producto)
    if req_kw and not df_f.empty:
        req_pattern = "|".join(req_kw)
        df_req = df_f[df_f["title_lower"].str.contains(req_pattern, na=False)]
        if not df_req.empty:
            df_f = df_req  # Solo si hay resultados; si no, mantenemos sin filtro

    if df_f.empty:
        return {"no_encontrado": True, "motivo": "Sin productos del tipo adecuado en Cecotec"}

    # 5. Filtrar más baratos que referencia
    if precio_ref > 0:
        df_cheap = df_f[df_f["precio_final"] < precio_ref].copy()
        if df_cheap.empty:
            df_cheap = df_f[df_f["precio_final"] <= precio_ref * 1.15].copy()
        df_f = df_cheap if not df_cheap.empty else df_f

    if df_f.empty:
        return {"no_encontrado": True, "motivo": "No hay alternativas Cecotec más económicas"}

    # 6. Score: solapamiento de palabras entre ref y título/desc Cecotec
    #    Bonus por rango de precio similar al competidor
    ref_words = set(re.findall(r'\w{4,}', texto_ref))

    # Subtype keywords: if ref mentions these, reward Cecotec products that also mention them
    SUBTYPE_PAIRS = [
        (["cepillo","cabezal","reversible","pelusas"], ["cepillo","hidrosteam","hydrosteam","vapor"]),
        (["vertical","colgar","colgante"], ["vertical","ironhero","hydrosteam"]),
        (["viaje","portatil","compacto","plegable"], ["viaje","folding","compacto"]),
        (["horizontal","suela","golpe vapor"], ["ironhero","plancha","suela"]),
        (["robot","autonomo","mapeado"], ["conga","robot"]),
        (["escoba","palo","inalambric"], ["rockstar","scoba","conga"]),
        (["freidora","air fryer","sin aceite"], ["cecofry","airfry","freido"]),
    ]

    def score_row(row):
        haystack = row["title_lower"] + " " + row["desc_lower"]
        kw_score = sum(1 for w in ref_words if w in haystack)
        # Subtype bonus
        for ref_kws, cec_kws in SUBTYPE_PAIRS:
            ref_match = any(k in texto_ref for k in ref_kws)
            cec_match = any(k in haystack for k in cec_kws)
            if ref_match and cec_match:
                kw_score += 3
            elif ref_match and not cec_match:
                kw_score -= 1  # slight penalty for subtype mismatch
        # Bonus: precio en rango 40-100% del precio de referencia
        p = row["precio_final"]
        price_bonus = 2 if (precio_ref * 0.4 <= p <= precio_ref) else 0
        return kw_score + price_bonus

    df_f = df_f.copy()
    df_f["_score"] = df_f.apply(score_row, axis=1)
    df_f = df_f.sort_values(["_score","precio_final"], ascending=[False, True])

    # 7. Seleccionar 3 alternativas: mejor match (recomendado), más barato, más caro
    best      = df_f.iloc[0]
    cheapest  = df_f.sort_values("precio_final").iloc[0]
    priciest  = df_f.sort_values("precio_final").iloc[-1]

    # Intermedio: el más cercano a la mediana de precios
    median_p  = df_f["precio_final"].median()
    mid_idx   = (df_f["precio_final"] - median_p).abs().idxmin()
    middle    = df_f.loc[mid_idx]

    # Evitar duplicados entre las 3 opciones
    seen_refs = set()
    alternatives = []
    for label, row in [("⭐ Recomendado", best), ("💰 Más económico", cheapest),
                        ("🎯 Intermedio", middle), ("🏆 Premium", priciest)]:
        ref_id = str(row.get("mpn","") or row["title"])
        if ref_id not in seen_refs:
            seen_refs.add(ref_id)
            alt = _make_alt(row, precio_ref, ref=ref)
            alt["label"] = label
            alternatives.append(alt)
        if len(alternatives) == 3:
            break

    main = alternatives[0]
    main["alternativas"] = alternatives
    return main

# ── Render resultados ─────────────────────────────────────────────────────────

def get_stock_row(ref_str: str, df_stk):
    if df_stk is None or df_stk.empty or not ref_str:
        return None
    r = df_stk[df_stk["Referencia"] == str(ref_str).strip()]
    return r.iloc[0] if not r.empty else None

def render_results(results, df_stk=None):
    rows, encontrados, ahorro_total = [], 0, 0.0
    for r in results:
        ref, alt = r["ref"], r["alt"]
        titulo = ref.get("titulo","")
        stk = get_stock_row(alt.get("cecotec_referencia",""), df_stk)
        st_op  = int(stk["Stock Operativo"]) if stk is not None else None
        st_mar = int(stk["Mar"])             if stk is not None else None
        st_pto = int(stk["Puerto"])          if stk is not None else None
        if alt.get("no_encontrado"):
            rows.append({
                "ASIN": ref.get("asin",""),
                "Producto competidor": titulo[:80],
                "Marca": ref.get("fabricante",""),
                "Precio comp. (€)": ref.get("precio"),
                "Subcategoría": ref.get("subcategoria",""),
                "Alternativa Cecotec": "❌ " + alt.get("motivo","")[:70],
                "Precio Cecotec (€)": None, "Ahorro (€)": None,
                "Prestaciones":"—","Ref. Cecotec":"—","URL Cecotec":"",
                "Conclusión": "—",
                "Stock Operativo": None, "Stock Mar": None, "Stock Puerto": None,
            })
        else:
            encontrados += 1
            ahorro = float(alt.get("ahorro_eur") or 0)
            ahorro_total += ahorro
            pmap = {"mejor":"✅ Mejor","igual":"🟡 Igual","peor":"🔴 Peor"}
            rows.append({
                "ASIN": ref.get("asin",""),
                "Producto competidor": titulo[:80],
                "Marca": ref.get("fabricante",""),
                "Precio comp. (€)": ref.get("precio"),
                "Subcategoría": ref.get("subcategoria",""),
                "Alternativa Cecotec": alt.get("cecotec_nombre",""),
                "Precio Cecotec (€)": alt.get("cecotec_precio"),
                "Ahorro (€)": round(ahorro, 2),
                "Prestaciones": pmap.get(alt.get("prestaciones",""), alt.get("prestaciones","")),
                "Ref. Cecotec": alt.get("cecotec_referencia",""),
                "URL Cecotec": alt.get("cecotec_url",""),
                "Conclusión": alt.get("justificacion",""),
                "Stock Operativo": st_op,
                "Stock Mar": st_mar,
                "Stock Puerto": st_pto,
            })

    df_res = pd.DataFrame(rows)
    st.markdown(f"""<div class="kpi-row">
      <div class="kpi"><div class="val">{encontrados}</div><div class="lbl">Alternativas encontradas</div></div>
      <div class="kpi"><div class="val">{len(results)-encontrados}</div><div class="lbl">Sin alternativa</div></div>
      <div class="kpi"><div class="val">{ahorro_total:.0f} €</div><div class="lbl">Ahorro total acumulado</div></div>
      <div class="kpi"><div class="val">{encontrados*100//len(results) if results else 0}%</div><div class="lbl">Tasa de cobertura</div></div>
    </div>""", unsafe_allow_html=True)

    st.dataframe(df_res, use_container_width=True, hide_index=True, column_config={
        "Precio comp. (€)":   st.column_config.NumberColumn(format="%.2f €"),
        "Precio Cecotec (€)": st.column_config.NumberColumn(format="%.2f €"),
        "Ahorro (€)":         st.column_config.NumberColumn(format="%.2f €"),
        "URL Cecotec":        st.column_config.LinkColumn("URL Cecotec"),
        "Conclusión":         st.column_config.TextColumn("Conclusión", width="large"),
        "Stock Operativo":    st.column_config.NumberColumn("Stock Disponible", format="%d uds"),
        "Stock Mar":          st.column_config.NumberColumn("Stock Mar", format="%d uds"),
        "Stock Puerto":       st.column_config.NumberColumn("Stock Puerto", format="%d uds"),
    })
    # ── Export buttons ──────────────────────────────────────────────────────
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    def build_excel(df: pd.DataFrame) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Comparativa Cecotec"

        # Brand colors
        BLK   = "FF141413"
        BLUE  = "FF3EB1C8"
        WHITE = "FFFFFFFF"
        LGREY = "FFF5F5F3"
        thin  = Side(style="thin", color="FFD0D0CC")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)

        # Header row
        headers = list(df.columns)
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font      = Font(bold=True, color=WHITE, name="Arial", size=9)
            cell.fill      = PatternFill("solid", start_color=BLK)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = border
        ws.row_dimensions[1].height = 30

        # Data rows
        for ri, row_data in enumerate(df.itertuples(index=False), 2):
            fill = PatternFill("solid", start_color=LGREY if ri % 2 == 0 else WHITE)
            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font      = Font(name="Arial", size=9)
                cell.fill      = fill
                cell.border    = border
                cell.alignment = Alignment(vertical="center")
                # Highlight stock columns: red if 0, green if >0
                col_name = headers[ci-1]
                if col_name in ("Stock Operativo","Stock Mar","Stock Puerto") and isinstance(val, (int, float)):
                    if val and val > 0:
                        cell.fill = PatternFill("solid", start_color="FFD4EDDA")
                        cell.font = Font(name="Arial", size=9, color="FF155724", bold=True)
                    elif val == 0:
                        cell.fill = PatternFill("solid", start_color="FFFCE4EC")
                        cell.font = Font(name="Arial", size=9, color="FFB71C1C")
                # URL as hyperlink
                if col_name == "URL Cecotec" and val and str(val).startswith("http"):
                    cell.hyperlink = str(val)
                    cell.value = "🔗 Ver"
                    cell.font = Font(name="Arial", size=9, color="FF3EB1C8", underline="single")

        # Column widths
        col_widths = {
            "ASIN": 14, "Producto competidor": 40, "Marca": 14,
            "Precio comp. (€)": 13, "Subcategoría": 22,
            "Alternativa Cecotec": 38, "Precio Cecotec (€)": 13,
            "Ahorro (€)": 10, "Prestaciones": 12, "Ref. Cecotec": 14,
            "URL Cecotec": 10, "Stock Operativo": 13, "Stock Mar": 11, "Stock Puerto": 12,
        }
        for ci, h in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(h, 15)

        # Freeze header
        ws.freeze_panes = "A2"

        # Logo row above header
        ws.insert_rows(1)
        ws.row_dimensions[1].height = 22
        logo_cell = ws.cell(row=1, column=1, value="cecotec · Comparador de Competencia")
        logo_cell.font = Font(bold=True, name="Arial", size=11, color=WHITE)
        logo_cell.fill = PatternFill("solid", start_color=BLK)
        logo_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def build_pdf(df: pd.DataFrame) -> bytes:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                leftMargin=10*mm, rightMargin=10*mm,
                                topMargin=12*mm, bottomMargin=10*mm)
        styles = getSampleStyleSheet()
        cec_blue  = colors.HexColor("#3EB1C8")
        cec_black = colors.HexColor("#141413")
        cec_bg    = colors.HexColor("#FAF9F5")

        title_style = ParagraphStyle("title", parent=styles["Heading1"],
                                     textColor=cec_black, fontSize=14, spaceAfter=4)
        sub_style   = ParagraphStyle("sub", parent=styles["Normal"],
                                     textColor=colors.HexColor("#6b7280"), fontSize=8, spaceAfter=8)

        elements = [
            Paragraph("cecotec · Comparador de Competencia", title_style),
            Paragraph(f"Exportado con {len(df)} productos comparados", sub_style),
            Spacer(1, 4*mm),
        ]

        # Select key columns for PDF (avoid too wide)
        pdf_cols = ["Producto competidor","Marca","Precio comp. (€)",
                    "Alternativa Cecotec","Precio Cecotec (€)","Ahorro (€)",
                    "Prestaciones","Stock Operativo","Stock Mar","Stock Puerto","URL Cecotec"]
        pdf_cols = [c for c in pdf_cols if c in df.columns]
        df_pdf = df[pdf_cols].copy()
        df_pdf["URL Cecotec"] = df_pdf["URL Cecotec"].apply(
            lambda x: "Ver →" if str(x).startswith("http") else x)

        col_widths_pdf = {
            "Producto competidor": 60*mm, "Marca": 22*mm,
            "Precio comp. (€)": 18*mm, "Alternativa Cecotec": 55*mm,
            "Precio Cecotec (€)": 18*mm, "Ahorro (€)": 16*mm,
            "Prestaciones": 18*mm, "Stock Operativo": 18*mm,
            "Stock Mar": 16*mm, "Stock Puerto": 18*mm, "URL Cecotec": 14*mm,
        }
        widths = [col_widths_pdf.get(c, 20*mm) for c in pdf_cols]

        data = [pdf_cols] + [[str(v) if v is not None else "—" for v in row]
                              for row in df_pdf.itertuples(index=False)]

        table = Table(data, colWidths=widths, repeatRows=1)
        style = TableStyle([
            ("BACKGROUND",  (0,0), (-1,0), cec_black),
            ("TEXTCOLOR",   (0,0), (-1,0), colors.white),
            ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",    (0,0), (-1,0), 7),
            ("ALIGN",       (0,0), (-1,0), "CENTER"),
            ("FONTNAME",    (0,1), (-1,-1), "Helvetica"),
            ("FONTSIZE",    (0,1), (-1,-1), 6.5),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F5F5F3")]),
            ("GRID",        (0,0), (-1,-1), 0.3, colors.HexColor("#D0D0CC")),
            ("VALIGN",      (0,0), (-1,-1), "MIDDLE"),
            ("TOPPADDING",  (0,0), (-1,-1), 3),
            ("BOTTOMPADDING",(0,0), (-1,-1), 3),
            ("LINEBELOW",   (0,0), (-1,0), 2, cec_blue),
        ])
        # Color stock cells
        for ri, row_data in enumerate(df_pdf.itertuples(index=False), 1):
            for ci, col in enumerate(pdf_cols):
                if col in ("Stock Operativo","Stock Mar","Stock Puerto"):
                    val = getattr(row_data, col.replace(" ","_"), None)
                    try:
                        v = int(val)
                        if v > 0:
                            style.add("BACKGROUND", (ci, ri), (ci, ri), colors.HexColor("#D4EDDA"))
                            style.add("TEXTCOLOR",  (ci, ri), (ci, ri), colors.HexColor("#155724"))
                        else:
                            style.add("BACKGROUND", (ci, ri), (ci, ri), colors.HexColor("#FCE4EC"))
                            style.add("TEXTCOLOR",  (ci, ri), (ci, ri), colors.HexColor("#B71C1C"))
                    except: pass
        table.setStyle(style)
        elements.append(table)
        doc.build(elements)
        return buf.getvalue()

    col_csv, col_xls, col_pdf = st.columns(3)
    with col_csv:
        st.download_button("⬇️ CSV", df_res.to_csv(index=False).encode("utf-8"),
                           "comparativa_cecotec.csv", "text/csv", use_container_width=True)
    with col_xls:
        try:
            st.download_button("⬇️ Excel", build_excel(df_res),
                               "comparativa_cecotec.xlsx",
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)
        except Exception as e:
            st.warning(f"Excel no disponible: {e}")
    with col_pdf:
        try:
            st.download_button("⬇️ PDF", build_pdf(df_res),
                               "comparativa_cecotec.pdf", "application/pdf",
                               use_container_width=True)
        except Exception as e:
            st.warning(f"PDF no disponible: {e}")

    found = [r for r in results if not r["alt"].get("no_encontrado")]
    if found:
        st.markdown("---")
        st.markdown('<div class="cec-section-title">🔎 Detalle por producto</div>', unsafe_allow_html=True)
        for r in found:
            ref, alt = r["ref"], r["alt"]
            alternatives = alt.get("alternativas", [alt])
            main_alt = alternatives[0]
            ahorro = float(main_alt.get("ahorro_eur") or 0)
            with st.expander(
                f"**{ref.get('titulo','')[:55]}** · {ref.get('precio','')}€  →  "
                f"**{main_alt.get('cecotec_nombre','')[:45]}** · {main_alt.get('cecotec_precio','')}€"
                + (f"  💰 -{ahorro:.2f}€" if ahorro > 0 else "")
            ):
                # Left: competitor product
                c1, mid, c2 = st.columns([4,1,7])
                with c1:
                    st.markdown("##### 📦 Producto competidor")
                    st.markdown(f"**{ref.get('titulo','')}**")
                    st.markdown(f"*{ref.get('fabricante','')}* · {ref.get('subcategoria','')}")
                    st.markdown(f"💶 **{ref.get('precio','')} €**")
                    st.markdown(f"_{ref.get('caracteristicas','')[:280]}_")
                    if ref.get("url_amazon"):
                        st.markdown(f"[🔗 Ver en Amazon]({ref['url_amazon']})")
                with mid:
                    st.markdown("<div style='font-size:2rem;text-align:center;margin-top:50px'>→</div>", unsafe_allow_html=True)
                with c2:
                    st.markdown("##### 🟦 Alternativas Cecotec")
                    # Tab per alternative
                    tab_labels = [a.get("label","Opción") for a in alternatives]
                    alt_tabs = st.tabs(tab_labels)
                    for tab, a in zip(alt_tabs, alternatives):
                        with tab:
                            prest = a.get("prestaciones","")
                            tag_css   = {"mejor":"tag-mejor","igual":"tag-igual","peor":"tag-peor"}.get(prest,"tag-skip")
                            tag_label = {"mejor":"✅ Mejor valor","igual":"🟡 Equivalente","peor":"🔴 Inferior"}.get(prest,prest)
                            st.markdown(f"**{a.get('cecotec_nombre','')}**")
                            p_orig = a.get("cecotec_precio_original")
                            if p_orig:
                                st.markdown(f"~~{p_orig}€~~ → 💶 **{a.get('cecotec_precio','')} €**")
                            else:
                                st.markdown(f"💶 **{a.get('cecotec_precio','')} €**")
                            st.markdown(f'<span class="{tag_css}">{tag_label}</span>', unsafe_allow_html=True)
                            ah = float(a.get("ahorro_eur") or 0)
                            if ah > 0:
                                st.markdown(f"💰 **Ahorro: {ah:.2f} €**")
                            # Conclusion
                            if a.get("justificacion"):
                                st.markdown(a["justificacion"])
                            st.markdown(f"_{a.get('cecotec_caracteristicas','')}_")
                            st.caption(f"📂 {a.get('cecotec_categoria','')}  ·  🏷️ Ref: {a.get('cecotec_referencia','')}")
                            if a.get("cecotec_url"):
                                st.markdown(f"[🔗 Ver en Cecotec.es]({a['cecotec_url']})")

def run_search_local(df_proc: pd.DataFrame, df_cec: pd.DataFrame) -> list:
    """Instant local matching — no API calls, no waiting."""
    results = []
    prog = st.progress(0)
    total = len(df_proc)
    for i, (_, row) in enumerate(df_proc.iterrows()):
        prog.progress((i+1)/total, text=f"[{i+1}/{total}] {str(row.get('titulo',''))[:60]}")
        alt = find_best_match_local(row.to_dict(), df_cec)
        results.append({"ref": row.to_dict(), "alt": alt})
    prog.empty()
    return results

# ── API key (solo para modo manual con ASIN) ─────────────────────────────────
def get_gemini_key():
    if st.session_state.get("GOOGLE_API_KEY"): return st.session_state["GOOGLE_API_KEY"]
    try:
        k = st.secrets.get("GOOGLE_API_KEY", None)
        if k: return k
    except Exception: pass
    return os.environ.get("GOOGLE_API_KEY", None)

# ── Load data ─────────────────────────────────────────────────────────────────
UPLOAD_DIR = Path(__file__).parent
_feed_path = st.session_state.get("feed_path")
if _feed_path and Path(_feed_path).exists():
    df_cecotec = load_cecotec_feed(_feed_path)
else:
    df_cecotec = load_cecotec_feed(str(UPLOAD_DIR))

df_keepa = load_keepa_files(str(UPLOAD_DIR))
_stock_path = st.session_state.get("stock_path")
if _stock_path and Path(_stock_path).exists():
    df_stock_data = load_stock(_stock_path)
else:
    df_stock_data = load_stock(str(UPLOAD_DIR))

# Status bar
feed_ok  = not df_cecotec.empty
keepa_ok = not df_keepa.empty
c1, c2, c3, c4, c5 = st.columns([3,3,2,2,1])
with c1:
    if feed_ok: st.success(f"✅ Feed Cecotec: **{len(df_cecotec):,}** productos en stock")
    else:       st.error("❌ Feed Cecotec no cargado")
with c2:
    if keepa_ok:
        df_rel = df_keepa[df_keepa["_cecotec_relevant"]]
        st.success(f"✅ Keepa: **{len(df_rel)}** relevantes de **{len(df_keepa)}** totales")
    else:
        st.warning("⚠️ Archivos Keepa no encontrados")
with c3:
    stock_ok = not df_stock_data.empty
    if stock_ok:
        st.success(f"✅ Stock: **{len(df_stock_data):,}** refs")
    else:
        st.warning("⚠️ Stock no cargado")
        stock_up = st.file_uploader("stock_cecotec.csv", type=["csv"], key="stock_up", label_visibility="collapsed")
        if stock_up:
            import tempfile
            tmp = Path(tempfile.mkdtemp()) / "stock_cecotec.csv"
            tmp.write_bytes(stock_up.read())
            st.session_state["stock_path"] = str(tmp)
            st.rerun()
with c4:
    st.info("⚡ **Matching local** · Sin IA")
with c5:
    if st.button("🗑️", help="Limpiar caché — fuerza recarga de todos los ficheros"):
        load_cecotec_feed.clear()
        load_keepa_files.clear()
        load_stock.clear()
        for k in ["results","results_custom","results_manual","feed_path","stock_path"]:
            st.session_state.pop(k, None)
        st.success("Caché limpiado")
        st.rerun()

# Feed uploader si no está cargado
if not feed_ok:
    st.markdown('<div class="cec-section-title">📂 Subir catálogo Cecotec</div>', unsafe_allow_html=True)
    st.info("Sube `feed_Espan_a.xlsx` para iniciar la app.")
    feed_upload = st.file_uploader("feed_Espan_a.xlsx", type=["xlsx","xls"], key="feed_upload")
    if feed_upload:
        import tempfile
        tmp = Path(tempfile.mkdtemp()) / "feed_Espan_a.xlsx"
        tmp.write_bytes(feed_upload.read())
        df_cecotec = load_cecotec_feed(str(tmp))
        st.session_state["feed_path"] = str(tmp)
        st.success(f"✅ {len(df_cecotec):,} productos cargados")
        st.rerun()
    else:
        st.stop()

# ── TABS ──────────────────────────────────────────────────────────────────────
# Auto-navigate to results tab if search just completed
_default_tab = 3 if st.session_state.pop("active_tab", None) == "resultados" else 0
if _default_tab == 3:
    st.toast("✅ Comparación completada — mostrando resultados", icon="🎯")

tab_keepa, tab_manual, tab_fichero, tab_resultados, tab_feed = st.tabs([
    "📦 Keepa Bestsellers","✏️ Producto manual",
    "📂 Subir fichero","📊 Resultados","🗄️ Feed Cecotec",
])

# ═══ TAB 1 · KEEPA ════════════════════════════════════════════════════════════
with tab_keepa:
    if not keepa_ok:
        st.warning("Archivos Keepa no encontrados.")
    else:
        df_relevant = df_keepa[df_keepa["_cecotec_relevant"]].copy()
        df_skipped  = df_keepa[~df_keepa["_cecotec_relevant"]].copy()
        st.markdown('<div class="cec-section-title">📊 Bestsellers Amazon · Análisis de competencia</div>', unsafe_allow_html=True)
        st.markdown(f"""<div class="kpi-row">
          <div class="kpi"><div class="val">{len(df_keepa[df_keepa['_source']=='Hogar']):,}</div><div class="lbl">Productos Hogar</div></div>
          <div class="kpi"><div class="val">{len(df_keepa[df_keepa['_source']=='Belleza']):,}</div><div class="lbl">Productos Belleza</div></div>
          <div class="kpi"><div class="val">{len(df_relevant)}</div><div class="lbl">Con equiv. Cecotec</div></div>
          <div class="kpi"><div class="val">{len(df_skipped)}</div><div class="lbl">Sin cobertura</div></div>
        </div>""", unsafe_allow_html=True)

        c1, c2, c3 = st.columns(3)
        with c1:
            src = st.multiselect("Origen", ["Hogar","Belleza"], default=["Hogar","Belleza"], key="k_src")
        with c2:
            pmax = float(df_relevant["precio"].max() or 500)
            pmin_v, pmax_v = st.slider("Precio competidor (€)", 0.0, pmax, (0.0, pmax), step=5.0, key="k_p")
        with c3:
            procesar_todos = st.checkbox("Procesar TODOS los relevantes", value=True)
            if not procesar_todos:
                n = st.number_input("Máx. productos", 1, len(df_relevant), min(50, len(df_relevant)), key="k_n")

        df_proc = df_relevant[
            df_relevant["_source"].isin(src) &
            df_relevant["precio"].between(pmin_v, pmax_v)
        ]
        if not procesar_todos:
            df_proc = df_proc.head(int(n))

        st.info(f"Se procesarán **{len(df_proc)}** productos · matching local instantáneo ⚡")
        if st.button("🚀 Comparar todos con Cecotec", type="primary", use_container_width=True, key="btn_k"):
            with st.spinner("Calculando…"):
                st.session_state["results"] = run_search_local(df_proc, df_cecotec)
            st.session_state["active_tab"] = "resultados"
            st.rerun()

# ═══ TAB 2 · MANUAL ═══════════════════════════════════════════════════════════
with tab_manual:
    st.markdown('<div class="cec-section-title">✏️ Buscar alternativa para un producto concreto</div>', unsafe_allow_html=True)
    st.caption("Introduce el ASIN — los datos se obtienen de Amazon automáticamente.")

    col_asin, col_price = st.columns([3,1])
    with col_asin:
        m_asin = st.text_input("ASIN Amazon *", placeholder="ej: B0BY9592V9")
    with col_price:
        m_precio_override = st.number_input("Precio (€) opcional", min_value=0.0, step=0.01, value=0.0)

    with st.expander("➕ Datos adicionales (opcional)"):
        cx1, cx2 = st.columns(2)
        with cx1:
            m_titulo = st.text_input("Nombre", placeholder="Se obtiene del ASIN")
            m_marca  = st.text_input("Marca", placeholder="Se obtiene del ASIN")
        with cx2:
            m_subcat = st.text_input("Categoría", placeholder="Se obtiene del ASIN")
            m_caract = st.text_area("Características", height=90, placeholder="Se obtienen del ASIN")

    if st.button("🔍 Buscar en Cecotec", type="primary", use_container_width=True, key="btn_m"):
        asin = m_asin.strip().upper()
        if not asin:
            st.error("El ASIN es obligatorio.")
        else:
            amz_data = {"titulo":m_titulo,"fabricante":m_marca,"subcategoria":m_subcat,
                        "caracteristicas":m_caract,"precio":m_precio_override or 0,
                        "asin":asin,"url_amazon":f"https://www.amazon.es/dp/{asin}","feat1":m_caract}
            with st.spinner(f"Obteniendo datos de Amazon para **{asin}**…"):
                try:
                    r = requests.get(f"https://www.amazon.es/dp/{asin}", headers=HEADERS, timeout=14)
                    soup = BeautifulSoup(r.text, "html.parser")
                    if not amz_data["titulo"]:
                        t = soup.select_one("#productTitle,#title")
                        if t: amz_data["titulo"] = t.get_text(strip=True)[:200]
                    if not amz_data["fabricante"]:
                        b = soup.select_one("#bylineInfo,.po-brand .a-span9")
                        if b: amz_data["fabricante"] = re.sub(r"Marca:|Visita la Store de","",b.get_text(strip=True)).strip()[:80]
                    if not amz_data["precio"]:
                        p = soup.select_one(".a-price .a-offscreen,#priceblock_ourprice")
                        if p:
                            try: amz_data["precio"] = float(re.sub(r"[^\d.]","", p.get_text().replace(",",".")))
                            except: pass
                    if not amz_data["caracteristicas"]:
                        bullets = soup.select("#feature-bullets .a-list-item")
                        amz_data["caracteristicas"] = " | ".join(b.get_text(strip=True) for b in bullets[:5])[:500]
                    if not amz_data["subcategoria"]:
                        crumbs = soup.select("#wayfinding-breadcrumbs_feature_div li")
                        if crumbs: amz_data["subcategoria"] = crumbs[-1].get_text(strip=True)
                    amz_data["feat1"] = amz_data["caracteristicas"]
                except Exception as e:
                    st.warning(f"Amazon no accesible: {e}")

            if amz_data.get("titulo"):
                st.success(f"✅ **{amz_data['titulo'][:60]}** · {amz_data.get('fabricante','')} · {amz_data.get('precio','')}€")
            if not amz_data["titulo"] and not amz_data["subcategoria"]:
                st.error("Sin datos. Rellena nombre y categoría manualmente.")
            else:
                with st.spinner("Buscando en feed Cecotec…"):
                    alt = find_best_match_local(amz_data, df_cecotec)
                    result = [{"ref": amz_data, "alt": alt}]
                    st.session_state["results_manual"] = result
                render_results(result, df_stk=df_stock_data)

    elif "results_manual" in st.session_state:
        render_results(st.session_state["results_manual"], df_stk=df_stock_data)

# ═══ TAB 3 · FICHERO ══════════════════════════════════════════════════════════
with tab_fichero:
    st.markdown('<div class="cec-section-title">📂 Subir fichero de productos competidores</div>', unsafe_allow_html=True)
    st.markdown("CSV o Excel con columnas: `titulo`, `fabricante`, `precio`, `subcategoria`, `caracteristicas` (o formato Keepa).")
    uploaded = st.file_uploader("CSV o Excel", type=["csv","xlsx","xls"], key="cfile")
    if uploaded:
        try:
            df_custom = load_custom_file(uploaded)
            df_rel_c  = df_custom[df_custom["_cecotec_relevant"]].copy()
            st.success(f"✅ {len(df_custom)} productos · {len(df_rel_c)} con equivalente Cecotec posible")
            st.dataframe(df_custom[["titulo","fabricante","precio","subcategoria","_cecotec_relevant"]].rename(
                columns={"_cecotec_relevant":"relevante"}), use_container_width=True, hide_index=True,
                column_config={"precio":st.column_config.NumberColumn(format="%.2f €")})
            ca, cb = st.columns(2)
            with ca: solo = st.checkbox("Solo con equiv. Cecotec posible", value=True)
            with cb: mx = st.number_input("Máx. productos", 1, len(df_custom), len(df_custom), key="mx_c")
            df_p = (df_rel_c if solo else df_custom).head(int(mx))
            st.info(f"Se procesarán **{len(df_p)}** productos · matching instantáneo ⚡")
            if st.button("🚀 Buscar alternativas", type="primary", use_container_width=True, key="btn_c"):
                with st.spinner("Calculando…"):
                    st.session_state["results_custom"] = run_search_local(df_p, df_cecotec)
                st.session_state["active_tab"] = "resultados"
                st.rerun()
        except Exception as e:
            st.error(f"Error: {e}")

# ═══ TAB 4 · RESULTADOS ═══════════════════════════════════════════════════════
with tab_resultados:
    res_key = None
    if st.session_state.get("results"):        res_key = "results"
    if st.session_state.get("results_custom"):
        if res_key:
            op = st.radio("Mostrar:", ["Keepa Bestsellers","Fichero propio"], horizontal=True)
            res_key = "results" if op == "Keepa Bestsellers" else "results_custom"
        else:
            res_key = "results_custom"
    if not res_key:
        st.info("Ejecuta una búsqueda en alguna de las pestañas anteriores.")
    else:
        render_results(st.session_state[res_key], df_stk=df_stock_data)

# ═══ TAB 5 · FEED ═════════════════════════════════════════════════════════════
with tab_feed:
    st.markdown('<div class="cec-section-title">🗄️ Catálogo Cecotec cargado</div>', unsafe_allow_html=True)
    st.markdown(f"""<div class="kpi-row">
      <div class="kpi"><div class="val">{len(df_cecotec):,}</div><div class="lbl">Productos en stock</div></div>
      <div class="kpi"><div class="val">{df_cecotec['categories'].nunique()}</div><div class="lbl">Categorías</div></div>
      <div class="kpi"><div class="val">{df_cecotec['precio_final'].min():.0f}–{df_cecotec['precio_final'].max():.0f} €</div><div class="lbl">Rango de precios</div></div>
      <div class="kpi"><div class="val">{(df_cecotec['sale_price'] != df_cecotec['price']).sum()}</div><div class="lbl">En oferta</div></div>
    </div>""", unsafe_allow_html=True)
    cat_f = st.multiselect("Filtrar categoría", sorted(df_cecotec["categories"].unique()), key="feed_cat")
    srch  = st.text_input("Buscar en título/descripción", key="feed_srch")
    df_show = df_cecotec.copy()
    if cat_f:  df_show = df_show[df_show["categories"].isin(cat_f)]
    if srch:   df_show = df_show[df_show["title"].str.contains(srch, case=False, na=False) | df_show["desc_clean"].str.contains(srch, case=False, na=False)]
    st.caption(f"{len(df_show):,} productos")
    st.dataframe(df_show[["title","categories","precio_final","price","link","desc_clean"]].rename(
        columns={"title":"Producto","categories":"Categoría","precio_final":"Precio (€)",
                 "price":"Precio orig.","link":"URL","desc_clean":"Descripción"}),
        use_container_width=True, hide_index=True,
        column_config={
            "Precio (€)":   st.column_config.NumberColumn(format="%.2f €"),
            "Precio orig.": st.column_config.NumberColumn(format="%.2f €"),
            "URL":          st.column_config.LinkColumn("URL"),
        })
